"""Genera Excel: 64 carte del mazzo con nome e score di flessibilità."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "carte-flessibilita.xlsx"

SIM_DECK_CODES = [
    118, 227, 238, 247, 328, 336, 348, 356, 367, 428, 437, 445, 456, 467, 478, 486,
    538, 548, 554, 564, 575, 577, 588, 586, 587, 637, 646, 655, 666, 663, 675, 678,
    674, 688, 687, 684, 747, 757, 758, 768, 766, 765, 776, 772, 773, 782, 784, 785,
    783, 846, 858, 857, 856, 868, 865, 864, 875, 874, 873, 877, 883, 885, 882, 881,
]

VALUES = ["Asso", "Due", "Tre", "Quattro", "Cinque", "Sei", "Sette", "Otto"]
SHAPES = ["Cerchi", "Cuori", "Triangoli", "Quadrati", "Stelle", "Esagoni", "Lampi", "Croci"]
COLORS = ["Rosso", "Arancio", "Giallo", "Verde", "Azzurro", "Blu", "Viola", "Bianco"]
POSITIONAL_COUNTS = [1, 3, 5, 7, 9, 11, 13, 15]
FEMININE_SHAPES = {"Stelle", "Croci"}
COLOR_FORMS = {
    "Bianco": {"m": "Bianco", "f": "Bianca", "mp": "Bianchi", "fp": "Bianche"},
    "Viola": {"fixed": "Viola"},
    "Blu": {"fixed": "Blu"},
    "Azzurro": {"m": "Azzurro", "f": "Azzurra", "mp": "Azzurri", "fp": "Azzurre"},
    "Verde": {"fixed": "Verde", "plural": "Verdi"},
    "Giallo": {"m": "Giallo", "f": "Gialla", "mp": "Gialli", "fp": "Gialle"},
    "Arancio": {"fixed": "Arancio"},
    "Rosso": {"m": "Rosso", "f": "Rossa", "mp": "Rossi", "fp": "Rosse"},
}
SCARCE_VALUES = {1, 2, 3}


def color_label(color_name, value, shape_name):
    forms = COLOR_FORMS[color_name]
    if "fixed" in forms:
        return forms["plural"] if value > 1 and "plural" in forms else forms["fixed"]
    feminine = shape_name in FEMININE_SHAPES
    if value > 1:
        return forms["fp"] if feminine else forms["mp"]
    return forms["f"] if feminine else forms["m"]


def format_card_name(code):
    text = str(code).zfill(3)
    value = int(text[0])
    shape = int(text[1])
    color = int(text[2])
    value_name = VALUES[value - 1]
    shape_name = SHAPES[shape - 1]
    color_name = COLORS[color - 1]
    color_word = color_label(color_name, value, shape_name)
    return f"{value_name} di {shape_name} {color_word}"


def compatibility_score(code):
    text = str(code).zfill(3)
    return sum(POSITIONAL_COUNTS[int(d) - 1] for d in text)


def rigidity_score(code):
    text = str(code).zfill(3)
    value = int(text[0])
    return 34 - compatibility_score(code) + (6 if value in SCARCE_VALUES else 0)


def decode(code):
    text = str(code).zfill(3)
    return int(text[0]), SHAPES[int(text[1]) - 1], COLORS[int(text[2]) - 1]


def build_cards():
    cards = []
    for code in SIM_DECK_CODES:
        val, shape, color = decode(code)
        flex = compatibility_score(code)
        cards.append({
            "codice": str(code).zfill(3),
            "nome": format_card_name(code),
            "valore": val,
            "forma": shape,
            "colore": color,
            "flessibilita": flex,
            "rigidita": rigidity_score(code),
        })
    return cards


def style_header(ws, cols):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True, name="Arial")
    for c in range(1, cols + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def flex_fill(score, min_s, max_s):
    if max_s == min_s:
        return PatternFill("solid", fgColor="FFFFFF")
    t = (score - min_s) / (max_s - min_s)
    if t <= 0.33:
        return PatternFill("solid", fgColor="FFC7CE")
    if t <= 0.66:
        return PatternFill("solid", fgColor="FFEB9C")
    return PatternFill("solid", fgColor="C6EFCE")


def write_sheet(ws, cards, with_rank=True):
    headers = [
        "Rank flessibilità",
        "Codice",
        "Nome",
        "Valore",
        "Forma",
        "Colore",
        "Score flessibilità",
        "Score rigidità",
        "Note",
    ]
    ws.append(headers)
    style_header(ws, len(headers))

    min_f = min(c["flessibilita"] for c in cards)
    max_f = max(c["flessibilita"] for c in cards)

    for i, card in enumerate(cards, 1):
        note = ""
        if card["valore"] in SCARCE_VALUES:
            note = "Valore raro (1–3)"
        if card["flessibilita"] == min_f:
            note = (note + "; " if note else "") + "Più rigida del mazzo"
        if card["flessibilita"] == max_f:
            note = (note + "; " if note else "") + "Più flessibile del mazzo"

        ws.append([
            i if with_rank else "",
            card["codice"],
            card["nome"],
            card["valore"],
            card["forma"],
            card["colore"],
            card["flessibilita"],
            card["rigidita"],
            note.strip("; "),
        ])

    widths = [14, 8, 32, 8, 12, 12, 16, 14, 28]
    for idx, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(idx)].width = w

    body_font = Font(name="Arial")
    for row in range(2, ws.max_row + 1):
        for col in range(1, len(headers) + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = body_font
            cell.alignment = Alignment(vertical="center", wrap_text=col == 3)
        flex_cell = ws.cell(row=row, column=7)
        flex_cell.fill = flex_fill(flex_cell.value, min_f, max_f)

    ws.freeze_panes = "A2"


def main():
    cards = build_cards()
    by_flex = sorted(cards, key=lambda c: (c["flessibilita"], c["codice"]))
    by_code = sorted(cards, key=lambda c: c["codice"])

    wb = Workbook()

    ws1 = wb.active
    ws1.title = "Per flessibilità"
    write_sheet(ws1, by_flex, with_rank=True)

    ws2 = wb.create_sheet("Per codice")
    write_sheet(ws2, by_code, with_rank=False)
    for row in range(2, ws2.max_row + 1):
        ws2.cell(row=row, column=1, value=row - 1)

    ws3 = wb.create_sheet("Legenda")
    ws3["A1"] = "Score di flessibilità (compatibilità nel mazzo)"
    ws3["A1"].font = Font(bold=True, name="Arial", size=12)
    lines = [
        "",
        "Per ogni carta si sommano i pesi delle tre cifre del codice (VALORE, FORMA, COLORE).",
        "Peso cifra k = numero di carte nel mazzo 64 che condividono quel tratto (1→1, 2→3, …, 8→15).",
        "",
        "Score alto = carta più «connessa» / flessibile (più partner potenziali nel mazzo).",
        "Score basso = carta più rigida / rara — da preferire su angoli e bordi.",
        "",
        "Score rigidità = 34 − flessibilità (+6 se valore 1, 2 o 3). Usato dal bot Durissima.",
        "",
        "Strategia M (Compatibilità): gioca per prima la carta con flessibilità minima tra le mosse legali.",
        "",
        f"Range nel mazzo 64: {min(c['flessibilita'] for c in cards)} – {max(c['flessibilita'] for c in cards)}.",
        "Codici: cifre 1–8 su ogni asse (da 111 a 888); il mazzo contiene 64 codici distinti.",
    ]
    for i, line in enumerate(lines, 2):
        ws3.cell(row=i, column=1, value=line).font = Font(name="Arial")
    ws3.column_dimensions["A"].width = 90

    wb.save(OUT)
    print(f"Scritto {OUT} ({len(cards)} carte)")


if __name__ == "__main__":
    main()