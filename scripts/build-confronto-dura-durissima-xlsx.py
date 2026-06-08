"""Excel comparativo: partita competitiva (Dura) vs Durissima Mater, tutte le L×G."""
import importlib.util
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "confronto-dura-durissima.xlsx"
TESTS = ROOT / "tests"


def load_durissima_results():
    path = ROOT / "scripts" / "build-durissima-riepilogo-xlsx.py"
    spec = importlib.util.spec_from_file_location("durissima_riepilogo", path)
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)
    return mod.RESULTS


def deal(size, players):
    total = size * size
    overcrowded = players > size
    cpp = total // players if overcrowded else size
    return cpp, total - cpp * players, overcrowded


def category(size, players):
    if players == 1:
        return "Solitario"
    if players == size:
        return "G=N (consigliato)"
    if players < size:
        return "Sotto-G"
    return "Overcrowd"


def playable(size, players):
    if players < 1 or players > 2 * size:
        return False
    cpp, _, _ = deal(size, players)
    return cpp >= 3


def extract_cells_from_doc(data):
    if data.get("cells"):
        return dict(data["cells"])
    cells = {}
    for step in data.get("steps", []):
        for k, v in (step.get("cells") or {}).items():
            cells[k] = v
    return cells


def load_classic_results(paths=None):
    if paths:
        files = [Path(p) for p in paths]
    else:
        files = sorted(TESTS.glob("dura-mater-classic-sweep-*.json"), key=lambda p: p.stat().st_mtime, reverse=True)
        seen = set()
        picked = []
        for p in files:
            stem = p.name
            if "L3-L5" in stem and "13-02-47" not in stem:
                continue
            tag = None
            for key in ("L3-L5", "L6", "L7", "L8"):
                if f"-{key}-" in stem or stem.endswith(f"-{key}-") or f"sweep-{key}-" in stem:
                    tag = key
                    break
            if tag is None:
                m = re.search(r"sweep-(L3-L5|L6|L7|L8)-", stem)
                tag = m.group(1) if m else stem
            if tag in seen:
                continue
            if (json.loads(p.read_text(encoding="utf-8")).get("countPerCell") or 0) < 500:
                continue
            seen.add(tag)
            picked.append(p)
        files = picked

    merged = {}
    for path in files:
        data = json.loads(path.read_text(encoding="utf-8"))
        for k, v in extract_cells_from_doc(data).items():
            merged[k] = v
    return merged, files


def classic_pct(raw):
    done = raw.get("done") or 0
    if not done:
        return None
    wins = raw.get("wins")
    if wins is None:
        wins = done - (raw.get("stalls") or 0)
    return wins / done


def durissima_pct(rec):
    n = rec.get("n") or 0
    if not n:
        return None
    return rec["ok"] / n


def interpretazione(dura, duri, cat):
    if dura is None or duri is None:
        return "Dati mancanti"
    delta = dura - duri
    if dura >= 0.9 and duri < 0.02:
        return "Dura ok · Durissima quasi impossibile"
    if dura >= 0.9 and duri >= 0.1:
        return "Entrambe giocabili"
    if dura < 0.8 and duri < 0.02:
        return "Entrambe difficili"
    if cat == "Solitario" and duri < 0.05:
        return "Solitario: vittoria possibile, griglia no"
    if delta >= 0.5:
        return "Dura molto più facile (obiettivo diverso)"
    if dura >= 0.95 and duri >= 0.05:
        return "Multigiocatore: entrambe con esiti"
    return "Dura competitiva più permissiva"


def build_rows(durissima_map, classic_map):
    rows = []
    for size in range(3, 9):
        for players in range(1, 2 * size + 1):
            if not playable(size, players):
                continue
            cpp, draw, _ = deal(size, players)
            cat = category(size, players)
            key = (size, players)
            drec = durissima_map.get(key)
            ckey = f"{size}x{players}"
            crec = classic_map.get(ckey)

            dura = classic_pct(crec) if crec else None
            duri = durissima_pct(drec) if drec else None
            delta = (dura - duri) if dura is not None and duri is not None else None

            rows.append({
                "N": size,
                "G": players,
                "formato": f"{size}×{players}",
                "carte_testa": cpp,
                "mazzo": draw,
                "categoria": cat,
                "dura_pct": dura,
                "dura_stalli": crec.get("stalls") if crec else None,
                "dura_turni": (crec.get("turnSum") or 0) / crec["done"] if crec and crec.get("done") else None,
                "duri_pct": duri,
                "duri_ok": drec["ok"] if drec else None,
                "duri_stalli": (drec["n"] - drec["ok"]) if drec else None,
                "delta": delta,
                "interpretazione": interpretazione(dura, duri, cat),
            })
    return rows


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True, name="Arial")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def fill_dura(cell, pct):
    if pct is None:
        cell.value = ""
        cell.fill = PatternFill("solid", fgColor="EDEDED")
    else:
        cell.value = pct
        cell.number_format = "0.0%"
        if pct >= 0.9:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif pct >= 0.7:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")


def fill_durissima(cell, pct):
    if pct is None:
        cell.value = ""
        cell.fill = PatternFill("solid", fgColor="EDEDED")
    else:
        cell.value = pct
        cell.number_format = "0.0%"
        if pct >= 0.1:
            cell.fill = PatternFill("solid", fgColor="C6EFCE")
        elif pct >= 0.02:
            cell.fill = PatternFill("solid", fgColor="FFEB9C")
        else:
            cell.fill = PatternFill("solid", fgColor="FFC7CE")


def fill_delta(cell, delta):
    if delta is None:
        cell.value = ""
        cell.fill = PatternFill("solid", fgColor="EDEDED")
    else:
        cell.value = delta
        cell.number_format = "0.0%"
        if delta >= 0.7:
            cell.fill = PatternFill("solid", fgColor="9DC3E6")
        elif delta >= 0.3:
            cell.fill = PatternFill("solid", fgColor="DDEBF7")
        else:
            cell.fill = PatternFill("solid", fgColor="FCE4D6")


def write_matrix(ws, title, rows, value_key, fill_fn, row_offset=1):
    ws.cell(row=row_offset, column=1, value=title).font = Font(bold=True, name="Arial")
    start = row_offset + 1
    ws.cell(row=start, column=1, value="N \\ G").font = Font(bold=True, name="Arial")
    max_g = 16
    for g in range(1, max_g + 1):
        ws.cell(row=start, column=g + 1, value=g).font = Font(bold=True, name="Arial")
        ws.cell(row=start, column=g + 1).alignment = Alignment(horizontal="center")

    val_map = {(r["N"], r["G"]): r.get(value_key) for r in rows}

    for i, size in enumerate(range(3, 9), start + 1):
        ws.cell(row=i, column=1, value=size).font = Font(bold=True, name="Arial")
        for g in range(1, max_g + 1):
            col = g + 1
            if not playable(size, g):
                cell = ws.cell(row=i, column=col, value="—")
                cell.fill = PatternFill("solid", fgColor="EDEDED")
            else:
                cell = ws.cell(row=i, column=col)
                fill_fn(cell, val_map.get((size, g)))
            cell.alignment = Alignment(horizontal="center")
            if playable(size, g) and g == size:
                cell.font = Font(bold=True, name="Arial")

    ws.column_dimensions["A"].width = 8
    for g in range(1, max_g + 1):
        ws.column_dimensions[get_column_letter(g + 1)].width = 6


def main():
    durissima_map = load_durissima_results()
    classic_map, classic_files = load_classic_results(
        sys.argv[1:] if len(sys.argv) > 1 else None
    )
    rows = build_rows(durissima_map, classic_map)

    wb = Workbook()

    ws = wb.active
    ws.title = "Confronto"
    headers = [
        "N", "G", "Formato", "Carte/testa", "Mazzo", "Categoria",
        "Dura Mater — Vittoria %", "Dura — Stalli", "Dura — Turni (med)",
        "Durissima — Griglia piena %", "Durissima — Completate", "Durissima — Stalli",
        "Δ (Dura − Durissima)", "Interpretazione",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r["N"], r["G"], r["formato"], r["carte_testa"], r["mazzo"], r["categoria"],
            r["dura_pct"] if r["dura_pct"] is not None else "",
            r["dura_stalli"] if r["dura_stalli"] is not None else "",
            r["dura_turni"] if r["dura_turni"] is not None else "",
            r["duri_pct"] if r["duri_pct"] is not None else "",
            r["duri_ok"] if r["duri_ok"] is not None else "",
            r["duri_stalli"] if r["duri_stalli"] is not None else "",
            r["delta"] if r["delta"] is not None else "",
            r["interpretazione"],
        ])

    for row in range(2, ws.max_row + 1):
        for col in (7, 10, 13):
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
        turns = ws.cell(row=row, column=9)
        if isinstance(turns.value, (int, float)):
            turns.number_format = "0.0"
        fill_dura(ws.cell(row=row, column=7), ws.cell(row=row, column=7).value or None)
        fill_durissima(ws.cell(row=row, column=10), ws.cell(row=row, column=10).value or None)
        fill_delta(ws.cell(row=row, column=13), ws.cell(row=row, column=13).value or None)

    widths = [5, 5, 10, 11, 11, 16, 18, 12, 14, 20, 18, 16, 16, 36]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wm = wb.create_sheet("Matrici")
    write_matrix(wm, "Dura Mater — % partite con vincitore (planner)", rows, "dura_pct", fill_dura, 1)
    write_matrix(wm, "Durissima Mater — % griglia N×N completata (durissima-planner)", rows, "duri_pct", fill_durissima, 12)
    write_matrix(
        wm,
        "Δ punti percentuali (Dura vittoria % − Durissima completamento %)",
        rows,
        "delta",
        fill_delta,
        23,
    )
    wm["A34"] = (
        "Legenda Dura: verde ≥90% · giallo 70–90% · rosso <70%  |  "
        "Durissima: verde ≥10% · giallo 2–10% · rosso <2%  |  "
        "Δ: blu = Dura molto più alta  |  grassetto = G=N"
    )
    wm["A34"].font = Font(italic=True, name="Arial", size=9)

    wi = wb.create_sheet("Info")
    info = [
        ["Confronto Dura Mater vs Durissima Mater", ""],
        ["", ""],
        ["Dura Mater (competitiva)", "Obiettivo: primo giocatore senza carte in mano"],
        ["Durissima Mater", "Obiettivo: riempire tutta la griglia N×N"],
        ["", ""],
        ["Dura — strategia", "planner su tutti i posti"],
        ["Durissima — strategia", "durissima-planner (isole + coop info mazzo)"],
        ["Partite per cella", "1000"],
        ["Configurazioni", "58 (G=1..2N, min 3 carte a testa)"],
        ["", ""],
        ["Export Dura", ", ".join(p.name for p in classic_files)],
        ["Export Durissima", "build-durissima-riepilogo-xlsx.py (RESULTS)"],
        ["", ""],
        ["Come leggere Δ", "Valori alti = competitiva finisce con vincitore molto più spesso del completamento griglia"],
        ["Solitario", "Spesso Dura ~75–100%, Durissima ~0% da ordine 5+"],
        ["Multigiocatore", "Dura quasi sempre ~100%; Durissima variabile ma >0% nella maggior parte dei casi"],
    ]
    for line in info:
        wi.append(line)
    wi["A1"].font = Font(bold=True, size=14, name="Arial")
    wi.column_dimensions["A"].width = 28
    wi.column_dimensions["B"].width = 72

    wb.save(OUT)
    missing_dura = sum(1 for r in rows if r["dura_pct"] is None)
    missing_duri = sum(1 for r in rows if r["duri_pct"] is None)
    print(f"Scritto: {OUT}")
    print(f"Righe: {len(rows)} · Dura mancanti: {missing_dura} · Durissima mancanti: {missing_duri}")


if __name__ == "__main__":
    main()