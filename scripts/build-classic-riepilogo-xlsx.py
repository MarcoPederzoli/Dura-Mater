"""Genera riepilogo Excel partita competitiva da export JSON (CLI o simulatore)."""
import json
import re
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "classic-riepilogo-test.xlsx"
TESTS = ROOT / "tests"


def deal(size, players):
    total = size * size
    overcrowded = players > size
    cpp = total // players if overcrowded else size
    return cpp, total - cpp * players, overcrowded


def category(size, players):
    if players == 1:
        return "Solitario"
    if players == size:
        return "G=N (consigliato)"
    if players < size:
        return "Sotto-G"
    return "Overcrowd"


def playable(size, players):
    if players < 1 or players > 2 * size:
        return False
    cpp, _, _ = deal(size, players)
    return cpp >= 3


def extract_cells_from_doc(data):
    if data.get("cells"):
        return dict(data["cells"])
    cells = {}
    for step in data.get("steps", []):
        for k, v in (step.get("cells") or {}).items():
            cells[k] = v
    return cells


def normalize_cell(key, raw):
    m = re.match(r"(\d+)x(\d+)", key)
    if not m:
        return None
    L, G = int(m.group(1)), int(m.group(2))
    done = raw.get("done") or 0
    if not done:
        return None

    wins = raw.get("wins")
    if wins is None:
        wins = done - (raw.get("stalls") or 0)

    return {
        "N": L,
        "G": G,
        "formato": f"{L}×{G}",
        "partite": done,
        "vittorie": wins,
        "stalli": raw.get("stalls", done - wins),
        "win_pct": wins / done,
        "dm_pct": (raw.get("dmClosedCount") or 0) / done,
        "grid_pct": (raw.get("boardCompleteCount") or 0) / done,
        "all_placed_pct": (raw.get("gamesAllPlayersPlaced") or 0) / done,
        "last_placed_pct": (raw.get("gamesLastPlayerPlaced") or 0) / done,
        "avg_turns": (raw.get("turnSum") or 0) / done,
        "avg_cards": (raw.get("totalPlacementsSum") or 0) / done / G,
        "carte_testa": raw.get("initialHandSize") or deal(L, G)[0],
        "mazzo": raw.get("initialDrawCount") if raw.get("initialDrawCount") is not None else deal(L, G)[1],
        "categoria": category(L, G),
    }


def load_results(paths):
    merged = {}
    for path in paths:
        data = json.loads(path.read_text(encoding="utf-8"))
        for k, v in extract_cells_from_doc(data).items():
            merged[k] = v
    return merged


def build_rows(cell_map):
    rows = []
    for size in range(3, 9):
        for players in range(1, 2 * size + 1):
            if not playable(size, players):
                continue
            cpp, draw, _ = deal(size, players)
            key = f"{size}x{players}"
            rec = cell_map.get(key)
            if rec:
                row = normalize_cell(key, rec)
            else:
                row = {
                    "N": size,
                    "G": players,
                    "formato": f"{size}×{players}",
                    "partite": None,
                    "vittorie": None,
                    "stalli": None,
                    "win_pct": None,
                    "dm_pct": None,
                    "grid_pct": None,
                    "all_placed_pct": None,
                    "last_placed_pct": None,
                    "avg_turns": None,
                    "avg_cards": None,
                    "carte_testa": cpp,
                    "mazzo": draw,
                    "categoria": category(size, players),
                    "note": "non testato",
                }
            if row:
                rows.append(row)
    return rows


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True, name="Arial")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    if len(sys.argv) > 1:
        paths = [Path(p) for p in sys.argv[1:]]
    else:
        paths = sorted(
            TESTS.glob("dura-mater-classic-sweep-*.json"),
            key=lambda p: p.stat().st_mtime,
            reverse=True,
        )
        if not paths:
            paths = sorted(
                TESTS.glob("dura-mater-sim-classic-audit-*.json"),
                key=lambda p: p.stat().st_mtime,
                reverse=True,
            )

    if not paths:
        print("Nessun export classic in tests/. Esegui prima classic-sweep.js o il workflow nel simulatore.")
        sys.exit(1)

    print("Fonti:")
    for p in paths:
        print(f"  - {p.name}")

    cell_map = load_results(paths)
    rows = build_rows(cell_map)
    tested = sum(1 for r in rows if r.get("partite"))

    wb = Workbook()

    ws = wb.active
    ws.title = "Dettaglio"
    headers = [
        "N", "G", "Formato", "Carte/testa", "Mazzo pesca", "Categoria",
        "Partite", "Vittorie", "Stalli", "Vittoria %", "DM chiusa %",
        "Griglia piena %", "Tutti posano %", "Ultimo posa %",
        "Turni (media)", "Carte posate/g", "Note"
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for r in rows:
        ws.append([
            r["N"], r["G"], r["formato"], r["carte_testa"], r["mazzo"], r["categoria"],
            r.get("partite"), r.get("vittorie"), r.get("stalli"),
            r.get("win_pct") if r.get("win_pct") is not None else "",
            r.get("dm_pct") if r.get("dm_pct") is not None else "",
            r.get("grid_pct") if r.get("grid_pct") is not None else "",
            r.get("all_placed_pct") if r.get("all_placed_pct") is not None else "",
            r.get("last_placed_pct") if r.get("last_placed_pct") is not None else "",
            r.get("avg_turns") if r.get("avg_turns") is not None else "",
            r.get("avg_cards") if r.get("avg_cards") is not None else "",
            r.get("note", ""),
        ])

    pct_cols = [10, 11, 12, 13, 14]
    for row in range(2, ws.max_row + 1):
        for col in pct_cols:
            cell = ws.cell(row=row, column=col)
            if isinstance(cell.value, (int, float)):
                cell.number_format = "0.0%"
        turns = ws.cell(row=row, column=15)
        if isinstance(turns.value, (int, float)):
            turns.number_format = "0.0"
        cards = ws.cell(row=row, column=16)
        if isinstance(cards.value, (int, float)):
            cards.number_format = "0.00"

    widths = [5, 5, 10, 11, 12, 16, 9, 10, 9, 11, 12, 13, 13, 13, 12, 13, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wm = wb.create_sheet("Matrice vittoria %")
    wm["A1"] = "N \\ G"
    wm["A1"].font = Font(bold=True, name="Arial")
    max_g = 16
    for g in range(1, max_g + 1):
        wm.cell(row=1, column=g + 1, value=g).font = Font(bold=True, name="Arial")
        wm.cell(row=1, column=g + 1).alignment = Alignment(horizontal="center")

    win_map = {(r["N"], r["G"]): r.get("win_pct") for r in rows if r.get("win_pct") is not None}
    grid_map = {(r["N"], r["G"]): r.get("grid_pct") for r in rows if r.get("grid_pct") is not None}

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="EDEDED")

    for i, size in enumerate(range(3, 9), 2):
        wm.cell(row=i, column=1, value=size).font = Font(bold=True, name="Arial")
        for g in range(1, max_g + 1):
            col = g + 1
            if not playable(size, g):
                cell = wm.cell(row=i, column=col, value="—")
                cell.fill = gray
            else:
                pct = win_map.get((size, g))
                if pct is None:
                    cell = wm.cell(row=i, column=col, value="")
                    cell.fill = gray
                else:
                    cell = wm.cell(row=i, column=col, value=pct)
                    cell.number_format = "0.0%"
                    if pct >= 0.90:
                        cell.fill = green
                    elif pct >= 0.70:
                        cell.fill = yellow
                    else:
                        cell.fill = red
            cell.alignment = Alignment(horizontal="center")
            if playable(size, g) and g == size:
                cell.font = Font(bold=True, name="Arial")

    wm.column_dimensions["A"].width = 8
    for g in range(1, max_g + 1):
        wm.column_dimensions[get_column_letter(g + 1)].width = 6
    wm["A10"] = "Legenda vittoria: verde ≥90% · giallo 70–90% · rosso <70% · grassetto = G=N"
    wm["A10"].font = Font(italic=True, name="Arial", size=9)

    wg = wb.create_sheet("Matrice griglia piena %")
    wg["A1"] = "N \\ G"
    wg["A1"].font = Font(bold=True, name="Arial")
    for g in range(1, max_g + 1):
        wg.cell(row=1, column=g + 1, value=g).font = Font(bold=True, name="Arial")
    for i, size in enumerate(range(3, 9), 2):
        wg.cell(row=i, column=1, value=size).font = Font(bold=True, name="Arial")
        for g in range(1, max_g + 1):
            col = g + 1
            if not playable(size, g):
                wg.cell(row=i, column=col, value="—").fill = gray
            else:
                pct = grid_map.get((size, g))
                cell = wg.cell(row=i, column=col, value=pct if pct is not None else "")
                if isinstance(pct, (int, float)):
                    cell.number_format = "0.0%"
                else:
                    cell.fill = gray
            wg.cell(row=i, column=col).alignment = Alignment(horizontal="center")
    wg["A10"] = "Griglia N×N piena a fine partita (competitiva: raro; utile vs Durissima)"
    wg["A10"].font = Font(italic=True, name="Arial", size=9)

    wi = wb.create_sheet("Info")
    info = [
        ["Riepilogo test partita competitiva", ""],
        ["", ""],
        ["Strategia", "planner su tutti i posti"],
        ["Modalità", "Competitiva — vincitore = primo senza carte"],
        ["Workflow", "classic-audit-L3-L5, L6, L7, L8 (simulatore o CLI)"],
        ["", ""],
        ["Celle testate", f"{tested} / 58 configurazioni legali"],
        ["Export usati", ", ".join(p.name for p in paths)],
        ["", ""],
        ["Confronto Durissima", "Usa colonna Griglia piena % vs successo % Durissima"],
        ["Solitario", "Incluso nei dati ma da analizzare a parte"],
    ]
    for line in info:
        wi.append(line)
    wi["A1"].font = Font(bold=True, size=14, name="Arial")
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"Scritto: {OUT} ({tested} celle con dati)")


if __name__ == "__main__":
    main()