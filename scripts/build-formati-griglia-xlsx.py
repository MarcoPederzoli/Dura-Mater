"""Genera tabella L x G: carte in mano / pesca, X se sconsigliato."""
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MIN_HAND = 3
MAX_G_TABLE = 20
OUT = r"D:\Grok\projects\Dura Mater\formati-griglia-giocatori.xlsx"


def deal(size, players):
    total = size * size
    hand = size if players <= size else total // players
    return hand, total - hand * players


def max_players(size):
    return min(21, 2 * size)


def recommended_max(size):
    return size


def playable(size, players):
    if players < 1 or players > MAX_G_TABLE:
        return False
    if players > max_players(size):
        return False
    return deal(size, players)[0] >= MIN_HAND


def recommended(size, players):
    return playable(size, players) and players <= recommended_max(size)


def main():
    wb = Workbook()
    ws = wb.active
    ws.title = "Formati"

    header_font = Font(name="Arial", bold=True)
    row_font = Font(name="Arial", bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    not_ok_fill = PatternFill("solid", fgColor="F3F4F6")
    disc_fill = PatternFill("solid", fgColor="FEF3C7")

    ws["A1"] = "Griglia \\ G"
    ws["A1"].font = header_font
    ws["A1"].alignment = center
    for g in range(1, MAX_G_TABLE + 1):
        c = g + 1
        cell = ws.cell(row=1, column=c, value=g)
        cell.font = header_font
        cell.alignment = center
        ws.column_dimensions[get_column_letter(c)].width = 7
    ws.column_dimensions["A"].width = 10

    for r, size in enumerate(range(3, 9), start=2):
        label = ws.cell(row=r, column=1, value=f"{size}x{size}")
        label.font = row_font
        label.alignment = center
        for g in range(1, MAX_G_TABLE + 1):
            cell = ws.cell(row=r, column=g + 1)
            cell.alignment = center
            if not playable(size, g):
                cell.value = "-"
                cell.fill = not_ok_fill
                continue
            hand, draw = deal(size, g)
            text = f"{hand}/{draw}"
            if not recommended(size, g):
                text += " X"
                cell.fill = disc_fill
            cell.value = text
            cell.font = Font(name="Arial", size=10)

    leg = wb.create_sheet("Legenda")
    leg["A1"] = "Legenda"
    leg["A1"].font = Font(name="Arial", bold=True, size=12)
    lines = [
        ("Cella", "Significato"),
        ("3/4", "3 carte in mano, 4 nel mazzo di pesca"),
        ("X", "Formato ammesso ma sconsigliato (G > N, overcrowd)"),
        ("-", "Non ammesso (G > 2N o meno di 3 carte a testa)"),
        ("Sfondo giallo", "Overcrowd: G > N"),
        ("", ""),
        ("Massimo G", "2N per griglia"),
        ("3x3", "6"),
        ("4x4", "8"),
        ("5x5", "10"),
        ("6x6", "12"),
        ("7x7", "14"),
        ("8x8", "16"),
        ("", ""),
        ("Regole", "Max G=2N. Consigliato G=N (senza mazzo). G<=N: N carte. G>N: floor(N^2/G) uguali. Min 3 carte."),
    ]
    for i, row in enumerate(lines, start=2):
        leg.cell(row=i, column=1, value=row[0])
        leg.cell(row=i, column=2, value=row[1])
        leg.cell(row=i, column=1).font = Font(name="Arial")
        leg.cell(row=i, column=2).font = Font(name="Arial")
    leg.column_dimensions["A"].width = 22
    leg.column_dimensions["B"].width = 55

    wb.save(OUT)
    print(OUT)


if __name__ == "__main__":
    main()