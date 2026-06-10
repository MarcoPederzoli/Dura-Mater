"""Confronto % successo Durissima per tutte le varianti di regole testate."""
from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
TESTS = ROOT / "tests"
OUT = ROOT / "confronto-varianti-durissima.xlsx"

# --- Riepilogo 2026-06-07 (riserva N, 1000 partite/cella) — da build-durissima-riepilogo-xlsx.py
RIEPILOGO_1000 = {
    (3, 1): {"ok": 235, "n": 1000},
    (3, 2): {"ok": 249, "n": 1000},
    (3, 3): {"ok": 186, "n": 1000},
    (4, 1): {"ok": 62, "n": 1000},
    (4, 2): {"ok": 160, "n": 1000},
    (4, 3): {"ok": 216, "n": 1000},
    (4, 4): {"ok": 162, "n": 1000},
    (4, 5): {"ok": 177, "n": 1000},
    (5, 1): {"ok": 0, "n": 1000},
    (5, 2): {"ok": 16, "n": 1000},
    (5, 3): {"ok": 58, "n": 1000},
    (5, 4): {"ok": 84, "n": 1000},
    (5, 5): {"ok": 26, "n": 1000},
    (5, 6): {"ok": 61, "n": 1000},
    (5, 7): {"ok": 44, "n": 1000},
    (5, 8): {"ok": 66, "n": 1000},
    (6, 1): {"ok": 0, "n": 1000},
    (6, 2): {"ok": 0, "n": 1000},
    (6, 3): {"ok": 18, "n": 1000},
    (6, 4): {"ok": 29, "n": 1000},
    (6, 5): {"ok": 51, "n": 1000},
    (6, 6): {"ok": 36, "n": 1000},
    (6, 7): {"ok": 31, "n": 1000},
    (6, 8): {"ok": 23, "n": 1000},
    (6, 9): {"ok": 25, "n": 1000},
    (6, 10): {"ok": 21, "n": 1000},
    (6, 11): {"ok": 20, "n": 1000},
    (6, 12): {"ok": 14, "n": 1000},
    (7, 1): {"ok": 0, "n": 1000},
    (7, 2): {"ok": 0, "n": 1000},
    (7, 3): {"ok": 0, "n": 1000},
    (7, 4): {"ok": 8, "n": 1000},
    (7, 5): {"ok": 23, "n": 1000},
    (7, 6): {"ok": 23, "n": 1000},
    (7, 7): {"ok": 14, "n": 1000},
    (7, 8): {"ok": 18, "n": 1000},
    (7, 9): {"ok": 7, "n": 1000},
    (7, 10): {"ok": 6, "n": 1000},
    (7, 11): {"ok": 13, "n": 1000},
    (7, 12): {"ok": 7, "n": 1000},
    (7, 13): {"ok": 9, "n": 1000},
    (7, 14): {"ok": 9, "n": 1000},
    (8, 1): {"ok": 0, "n": 1000},
    (8, 2): {"ok": 0, "n": 1000},
    (8, 3): {"ok": 0, "n": 1000},
    (8, 4): {"ok": 1, "n": 1000},
    (8, 5): {"ok": 3, "n": 1000},
    (8, 6): {"ok": 8, "n": 1000},
    (8, 7): {"ok": 5, "n": 1000},
    (8, 8): {"ok": 17, "n": 1000},
    (8, 9): {"ok": 10, "n": 1000},
    (8, 10): {"ok": 5, "n": 1000},
    (8, 11): {"ok": 1, "n": 1000},
    (8, 12): {"ok": 6, "n": 1000},
    (8, 13): {"ok": 4, "n": 1000},
    (8, 14): {"ok": 3, "n": 1000},
    (8, 15): {"ok": 3, "n": 1000},
    (8, 16): {"ok": 2, "n": 1000},
}

# Pool N condiviso — sweep manuale 1000 partite (solo L3–L7)
POOL_N_1000 = {
    (3, 1): {"ok": 241, "n": 1000},
    (4, 1): {"ok": 128, "n": 1000},
    (5, 1): {"ok": 38, "n": 1000},
    (6, 1): {"ok": 7, "n": 1000},
    (7, 1): {"ok": 2, "n": 1000},
}

# Vita 1/giocatore — sweep manuale 1000 partite solitario (pre-pool)
VITA1_1000 = {
    (3, 1): {"ok": 221, "n": 1000},
    (4, 1): {"ok": 111, "n": 1000},
    (5, 1): {"ok": 7, "n": 1000},
    (6, 1): {"ok": 0, "n": 1000},
}

VARIANTS = [
    {
        "id": "normale",
        "name": "Durissima senza aiuto reattivo",
        "short": "Senza aiuto",
        "desc": "Solitario: nessun buffer emergenza né vita extra (budget pesca X=0). Solo L3–L8.",
        "bot": "planner (sweep budget)",
        "source": "durissima-budget-solitario 2026-06-07",
        "scope": "solo_only",
    },
    {
        "id": "buffer_xn",
        "name": "Buffer pesca emergenza X=N (solo)",
        "short": "Buffer X=N",
        "desc": "Solitario: fino a N pescate singole da fermo + pesca fine turno (budget X=N). Solo dove testato.",
        "bot": "planner",
        "source": "durissima-budget-solitario 2026-06-07",
        "scope": "solo_only",
    },
    {
        "id": "buffer",
        "name": "Buffer emergenza N pescate (default engine)",
        "short": "Buffer N",
        "desc": "Prima della vita extra: pescate da fermo in solitario (emergSum>0), multi senza buffer.",
        "bot": "durissima-planner",
        "source": "rules-probe 2026-06-07 (300/cella)",
        "scope": "partial",
    },
    {
        "id": "riserva",
        "name": "Riserva N + regole coop",
        "short": "Riserva N",
        "desc": "Riserva scoperta di N carte dal tallone; probe completo G=1..2N.",
        "bot": "durissima-planner",
        "source": "durissima-riepilogo 2026-06-07 (1000/cella)",
        "scope": "full",
    },
    {
        "id": "vita1",
        "name": "Vita extra 1 per giocatore",
        "short": "Vita 1/G",
        "desc": "1 reshuffle per giocatore per partita; multi facoltativo.",
        "bot": "durissima-planner",
        "source": "rules-probe 2026-06-09 (300/cella) + solo sweep 1000 L3–L6",
        "scope": "partial_plus_solo1000",
    },
    {
        "id": "pool",
        "name": "Pool vite extra N condiviso",
        "short": "Pool N",
        "desc": "Pool di N reshuffle condivisi; catena automatica in solitario.",
        "bot": "durissima-planner",
        "source": "engine attuale — solo sweep 1000 L3–L7",
        "scope": "solo_partial",
    },
]

FILL_HEADER = PatternFill("solid", fgColor="D9E1F2")
FILL_MISSING = PatternFill("solid", fgColor="FFF2CC")
FILL_NA = PatternFill("solid", fgColor="EDEDED")
FILL_GREEN = PatternFill("solid", fgColor="C6EFCE")
FILL_YELLOW = PatternFill("solid", fgColor="FFEB9C")
FILL_RED = PatternFill("solid", fgColor="FFC7CE")
FONT_BOLD = Font(bold=True, name="Arial")
FONT_ITALIC = Font(italic=True, name="Arial", size=9)


def deal(size: int, players: int):
    total = size * size
    overcrowded = players > size
    cpp = total // players if overcrowded else size
    return cpp, total - cpp, overcrowded


def playable(size: int, players: int) -> bool:
    if players < 1 or players > 2 * size:
        return False
    cpp, _, _ = deal(size, players)
    return cpp >= 3


def category(size: int, players: int) -> str:
    if players == 1:
        return "Solitario"
    if players == size:
        return "G=N"
    if players < size:
        return "Sotto-G"
    return "Overcrowd"


def pct_from_cell(cell: dict) -> tuple[float | None, int | None]:
    done = cell.get("done") or 0
    if not done:
        return None, None
    ok = done - (cell.get("stalls") or 0)
    return ok / done, done


def load_probe_cells(path: Path) -> dict[tuple[int, int], tuple[float, int]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    n = data.get("countPerCell")
    out: dict[tuple[int, int], tuple[float, int]] = {}
    cells = dict(data.get("cells") or {})
    for step in data.get("steps") or []:
        cells.update(step.get("cells") or {})
    for key, cell in cells.items():
        m = re.match(r"(\d+)x(\d+)", key)
        if not m:
            continue
        pct, done = pct_from_cell(cell)
        if pct is None:
            continue
        out[(int(m.group(1)), int(m.group(2)))] = (pct, done or n or 0)
    return out


def load_budget_solitario(path: Path) -> dict[str, dict[tuple[int, int], tuple[float, int]]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    out: dict[str, dict[tuple[int, int], tuple[float, int]]] = {"x0": {}, "xn": {}}
    for step in data.get("steps") or []:
        bx = step.get("budgetX")
        for key, cell in (step.get("cells") or {}).items():
            m = re.match(r"(\d+)x1", key)
            if not m:
                continue
            L = int(m.group(1))
            pct, done = pct_from_cell(cell)
            if pct is None:
                continue
            entry = (pct, done)
            if bx == 0:
                out["x0"][(L, 1)] = entry
            if bx == L:
                out["xn"][(L, 1)] = entry
    return out


def dict_pct(data: dict, key: tuple[int, int]) -> tuple[float | None, int | None, bool]:
    rec = data.get(key)
    if not rec:
        return None, None, False
    ok, n = rec["ok"], rec["n"]
    return ok / n, n, True


def all_configs():
    rows = []
    for size in range(3, 9):
        for players in range(1, 2 * size + 1):
            if not playable(size, players):
                continue
            cpp, draw, oc = deal(size, players)
            rows.append({
                "N": size,
                "G": players,
                "key": (size, players),
                "formato": f"{size}×{players}",
                "carte_testa": cpp,
                "mazzo": draw,
                "categoria": category(size, players),
                "overcrowd": oc,
            })
    return rows


def pct_fill(pct: float | None):
    if pct is None:
        return FILL_MISSING
    if pct >= 0.10:
        return FILL_GREEN
    if pct >= 0.02:
        return FILL_YELLOW
    return FILL_RED


def style_header_row(ws, ncol: int):
    for c in range(1, ncol + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = FILL_HEADER
        cell.font = FONT_BOLD
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_dataset():
    budget_path = TESTS / "dura-mater-durissima-budget-solitario-2026-06-07-00-30-26.json"
    buffer_path = TESTS / "dura-mater-durissima-rules-probe-2026-06-07-09-57-51.json"
    vita1_path = TESTS / "dura-mater-durissima-rules-probe-2026-06-09-13-25-32.json"

    budget = load_budget_solitario(budget_path) if budget_path.exists() else {"x0": {}, "xn": {}}
    buffer = load_probe_cells(buffer_path) if buffer_path.exists() else {}
    vita1_probe = load_probe_cells(vita1_path) if vita1_path.exists() else {}

    def get(variant_id: str, key: tuple[int, int]) -> tuple[float | None, int | None, str]:
        L, G = key
        if variant_id == "normale":
            if G != 1:
                return None, None, "solo"
            v = budget["x0"].get(key)
            return (v[0], v[1], "") if v else (None, None, "")
        if variant_id == "buffer_xn":
            if G != 1:
                return None, None, "solo"
            v = budget["xn"].get(key)
            return (v[0], v[1], "") if v else (None, None, "")
        if variant_id == "buffer":
            v = buffer.get(key)
            return (v[0], v[1], "") if v else (None, None, "")
        if variant_id == "riserva":
            p, n, ok = dict_pct(RIEPILOGO_1000, key)
            return (p, n, "") if ok else (None, None, "")
        if variant_id == "vita1":
            if key in vita1_probe:
                return vita1_probe[key][0], vita1_probe[key][1], ""
            p, n, ok = dict_pct(VITA1_1000, key)
            if ok:
                return p, n, "solo1000"
            return None, None, ""
        if variant_id == "pool":
            p, n, ok = dict_pct(POOL_N_1000, key)
            if ok:
                return p, n, ""
            return None, None, ""
        return None, None, ""

    return get


def main():
    configs = all_configs()
    get = build_dataset()

    wb = Workbook()

    # --- Legenda ---
    wl = wb.active
    wl.title = "Legenda"
    wl.append(["Variante", "Nome breve", "Descrizione", "Bot", "Fonte dati", "Copertura"])
    style_header_row(wl, 6)
    for v in VARIANTS:
        wl.append([v["name"], v["short"], v["desc"], v["bot"], v["source"], v["scope"]])
    wl.column_dimensions["A"].width = 34
    wl.column_dimensions["B"].width = 12
    wl.column_dimensions["C"].width = 58
    wl.column_dimensions["D"].width = 18
    wl.column_dimensions["E"].width = 36
    wl.column_dimensions["F"].width = 18
    wl.append([])
    wl.append(["Note generali", ""])
    notes = [
        "Bot durissima-planner salvo varianti budget (planner).",
        "Celle gialle = dato mancante (non simulato). Grigio = configurazione non ammessa (G>2N o <3 carte/testa).",
        "Riserva N: unica variante con matrice completa 1000 partite/cella.",
        "Pool N: regole engine attuali; multi non ancora simulato.",
        "Vita 1/G: probe 300/cella su subset; solitario L3–L6 anche a 1000 partite.",
    ]
    for n in notes:
        wl.append([n])

    # --- Confronto lungo ---
    wc = wb.create_sheet("Confronto")
    headers = [
        "N", "G", "Formato", "Categoria", "Carte/testa", "Mazzo",
    ]
    for v in VARIANTS:
        headers.append(f"{v['short']} %")
        headers.append(f"{v['short']} n")
    headers.append("Note celle")
    wc.append(headers)
    style_header_row(wc, len(headers))

    missing_log = []

    for row_i, cfg in enumerate(configs, 2):
        key = cfg["key"]
        note_parts = []
        row = [
            cfg["N"], cfg["G"], cfg["formato"], cfg["categoria"],
            cfg["carte_testa"], cfg["mazzo"],
        ]
        col = 7
        for v in VARIANTS:
            pct, n, tag = get(v["id"], key)
            row.append(pct if pct is not None else "")
            row.append(n if n is not None else "")
            cell_pct = wc.cell(row=row_i, column=col)
            if pct is None:
                cell_pct.fill = FILL_MISSING
                if v["scope"] == "solo_only" and cfg["G"] != 1:
                    cell_pct.value = "—"
                    cell_pct.fill = FILL_NA
                else:
                    missing_log.append((cfg["formato"], v["short"], tag or "mancante"))
            else:
                cell_pct.number_format = "0.0%"
                cell_pct.fill = pct_fill(pct)
                if tag:
                    note_parts.append(f"{v['short']}:{tag}")
            col += 2
        row.append("; ".join(note_parts))
        wc.append(row)

    for i, w in enumerate([5, 5, 10, 12, 10, 10] + [9, 6] * len(VARIANTS) + [24], 1):
        wc.column_dimensions[get_column_letter(i)].width = w

    # --- Matrici per variante ---
    max_g = 16
    for v in VARIANTS:
        sheet_title = re.sub(r'[\\/*?:\[\]]', " ", v["short"])[:31]
        ws = wb.create_sheet(sheet_title)
        ws["A1"] = "N \\ G"
        ws["A1"].font = FONT_BOLD
        for g in range(1, max_g + 1):
            ws.cell(row=1, column=g + 1, value=g).font = FONT_BOLD
        for i, size in enumerate(range(3, 9), 2):
            ws.cell(row=i, column=1, value=size).font = FONT_BOLD
            for g in range(1, max_g + 1):
                col = g + 1
                if not playable(size, g):
                    cell = ws.cell(row=i, column=col, value="—")
                    cell.fill = FILL_NA
                else:
                    pct, n, tag = get(v["id"], (size, g))
                    if pct is None:
                        cell = ws.cell(row=i, column=col, value="")
                        cell.fill = FILL_MISSING
                    else:
                        cell = ws.cell(row=i, column=col, value=pct)
                        cell.number_format = "0.0%"
                        cell.fill = pct_fill(pct)
                        if g == size:
                            cell.font = FONT_BOLD
                cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions["A"].width = 8
        for g in range(1, max_g + 1):
            ws.column_dimensions[get_column_letter(g + 1)].width = 6

    # --- Mancanti ---
    wm = wb.create_sheet("Celle mancanti")
    wm.append(["Formato", "Variante", "Motivo"])
    style_header_row(wm, 3)
    seen = set()
    for fmt, var, reason in sorted(set(missing_log)):
        if (fmt, var) in seen:
            continue
        seen.add((fmt, var))
        wm.append([fmt, var, reason or "non simulato"])
    wm.column_dimensions["A"].width = 10
    wm.column_dimensions["B"].width = 14
    wm.column_dimensions["C"].width = 24

    wb.save(OUT)
    print(f"Scritto: {OUT}")
    print(f"Configurazioni: {len(configs)} · celle mancanti (coppie): {len(seen)}")


if __name__ == "__main__":
    main()