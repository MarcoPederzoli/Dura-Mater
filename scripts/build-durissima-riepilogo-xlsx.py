"""Genera riepilogo Excel test Durissima cooperativo/solitario."""
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "durissima-riepilogo-test.xlsx"

# Bot: durissima-planner (isole + mazzo condiviso coop). 1000 partite/cella salvo nota.
RESULTS = {
    (3, 1): {"ok": 235, "n": 1000},
    (3, 2): {"ok": 249, "n": 1000},
    (3, 3): {"ok": 186, "n": 1000},
    (4, 1): {"ok": 62, "n": 1000},
    (4, 2): {"ok": 160, "n": 1000},
    (4, 3): {"ok": 216, "n": 1000},
    (4, 4): {"ok": 162, "n": 1000},
    (4, 5): {"ok": 177, "n": 1000},
    (5, 1): {"ok": 0, "n": 1000},
    (5, 2): {"ok": 16, "n": 1000},
    (5, 3): {"ok": 58, "n": 1000},
    (5, 4): {"ok": 84, "n": 1000},
    (5, 5): {"ok": 26, "n": 1000},
    (5, 6): {"ok": 61, "n": 1000},
    (5, 7): {"ok": 44, "n": 1000},
    (5, 8): {"ok": 66, "n": 1000},
    (6, 1): {"ok": 0, "n": 1000},
    (6, 2): {"ok": 0, "n": 1000},
    (6, 3): {"ok": 18, "n": 1000},
    (6, 4): {"ok": 29, "n": 1000},
    (6, 5): {"ok": 51, "n": 1000},
    (6, 6): {"ok": 36, "n": 1000},
    (6, 7): {"ok": 31, "n": 1000},
    (6, 8): {"ok": 23, "n": 1000},
    (6, 9): {"ok": 25, "n": 1000},
    (6, 10): {"ok": 21, "n": 1000},
    (6, 11): {"ok": 20, "n": 1000},
    (6, 12): {"ok": 14, "n": 1000},
    (7, 1): {"ok": 0, "n": 1000},
    (7, 2): {"ok": 0, "n": 1000},
    (7, 3): {"ok": 0, "n": 1000},
    (7, 4): {"ok": 8, "n": 1000},
    (7, 5): {"ok": 23, "n": 1000},
    (7, 6): {"ok": 23, "n": 1000},
    (7, 7): {"ok": 14, "n": 1000},
    (7, 8): {"ok": 18, "n": 1000},
    (7, 9): {"ok": 7, "n": 1000},
    (7, 10): {"ok": 6, "n": 1000},
    (7, 11): {"ok": 13, "n": 1000},
    (7, 12): {"ok": 7, "n": 1000},
    (7, 13): {"ok": 9, "n": 1000},
    (7, 14): {"ok": 9, "n": 1000},
    (8, 1): {"ok": 0, "n": 1000},
    (8, 2): {"ok": 0, "n": 1000},
    (8, 3): {"ok": 0, "n": 1000},
    (8, 4): {"ok": 1, "n": 1000},
    (8, 5): {"ok": 3, "n": 1000},
    (8, 6): {"ok": 8, "n": 1000},
    (8, 7): {"ok": 5, "n": 1000},
    (8, 8): {"ok": 17, "n": 1000},
    (8, 9): {"ok": 10, "n": 1000},
    (8, 10): {"ok": 5, "n": 1000},
    (8, 11): {"ok": 1, "n": 1000},
    (8, 12): {"ok": 6, "n": 1000},
    (8, 13): {"ok": 4, "n": 1000},
    (8, 14): {"ok": 3, "n": 1000},
    (8, 15): {"ok": 3, "n": 1000},
    (8, 16): {"ok": 2, "n": 1000},
}


def deal(size, players):
    total = size * size
    overcrowded = players > size
    cpp = total // players if overcrowded else size
    dealt = cpp * players
    return cpp, total - dealt, overcrowded


def category(size, players):
    if players == 1:
        return "Solitario"
    if players == size:
        return "G=N (consigliato)"
    if players < size:
        return "Sotto-G"
    return "Overcrowd"


def playable(size, players):
    if players < 1 or players > 2 * size:
        return False
    cpp, _, _ = deal(size, players)
    return cpp >= 3


def build_rows():
    rows = []
    for size in range(3, 9):
        for players in range(1, 2 * size + 1):
            if not playable(size, players):
                continue
            cpp, draw, _ = deal(size, players)
            key = (size, players)
            rec = RESULTS.get(key)
            if rec:
                n = rec["n"]
                ok = rec["ok"]
                pct = ok / n if n else None
                stalls = n - ok
                note = rec.get("note", "")
            else:
                n = ok = stalls = None
                pct = None
                note = "non testato"
            rows.append({
                "N": size,
                "G": players,
                "formato": f"{size}×{players}",
                "carte_testa": cpp,
                "mazzo": draw,
                "categoria": category(size, players),
                "partite": n,
                "ok": ok,
                "stalli": stalls,
                "pct": pct,
                "note": note,
            })
    return rows


def style_header(ws, row, cols):
    fill = PatternFill("solid", fgColor="D9E1F2")
    font = Font(bold=True, name="Arial")
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.font = font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def main():
    rows = build_rows()
    wb = Workbook()

    # --- Foglio Dettaglio ---
    ws = wb.active
    ws.title = "Dettaglio"
    headers = [
        "N", "G", "Formato", "Carte/testa", "Mazzo pesca", "Categoria",
        "Partite", "Completate", "Stalli", "Successo %", "Note"
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for r in rows:
        pct_val = r["pct"]
        ws.append([
            r["N"], r["G"], r["formato"], r["carte_testa"], r["mazzo"], r["categoria"],
            r["partite"], r["ok"], r["stalli"],
            pct_val if pct_val is not None else "",
            r["note"],
        ])

    for row in range(2, ws.max_row + 1):
        cell = ws.cell(row=row, column=10)
        if isinstance(cell.value, (int, float)):
            cell.number_format = "0.0%"

    widths = [5, 5, 10, 11, 12, 16, 9, 12, 9, 12, 28]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # --- Foglio Matrice % ---
    wm = wb.create_sheet("Matrice successo %")
    wm["A1"] = "N \\ G"
    wm["A1"].font = Font(bold=True, name="Arial")
    max_g = 16
    for g in range(1, max_g + 1):
        wm.cell(row=1, column=g + 1, value=g)
        wm.cell(row=1, column=g + 1).font = Font(bold=True, name="Arial")
        wm.cell(row=1, column=g + 1).alignment = Alignment(horizontal="center")

    pct_map = {(r["N"], r["G"]): r["pct"] for r in rows if r["pct"] is not None}
    cat_map = {(r["N"], r["G"]): r["categoria"] for r in rows}

    green = PatternFill("solid", fgColor="C6EFCE")
    yellow = PatternFill("solid", fgColor="FFEB9C")
    red = PatternFill("solid", fgColor="FFC7CE")
    gray = PatternFill("solid", fgColor="EDEDED")

    for i, size in enumerate(range(3, 9), 2):
        wm.cell(row=i, column=1, value=size)
        wm.cell(row=i, column=1).font = Font(bold=True, name="Arial")
        for g in range(1, max_g + 1):
            col = g + 1
            if not playable(size, g):
                cell = wm.cell(row=i, column=col, value="—")
                cell.fill = gray
            else:
                pct = pct_map.get((size, g))
                if pct is None:
                    cell = wm.cell(row=i, column=col, value="")
                    cell.fill = gray
                else:
                    cell = wm.cell(row=i, column=col, value=pct)
                    cell.number_format = "0.0%"
                    if pct >= 0.10:
                        cell.fill = green
                    elif pct >= 0.02:
                        cell.fill = yellow
                    else:
                        cell.fill = red
            cell.alignment = Alignment(horizontal="center")
            if playable(size, g) and cat_map.get((size, g)) == "G=N (consigliato)":
                cell.font = Font(bold=True, name="Arial")

    wm.column_dimensions["A"].width = 8
    for g in range(1, max_g + 1):
        wm.column_dimensions[get_column_letter(g + 1)].width = 6

    wm["A10"] = "Legenda: verde ≥10% · giallo 2–10% · rosso <2% · grigio non ammesso/non testato · grassetto = G=N"
    wm["A10"].font = Font(italic=True, name="Arial", size=9)

    # --- Foglio Info ---
    wi = wb.create_sheet("Info")
    info = [
        ["Riepilogo test Durissima Mater", ""],
        ["", ""],
        ["Bot", "durissima-planner (isole 2×2 + compattazione + info mazzo condiviso in coop)"],
        ["Modalità", "Durissima Mater — completamento griglia N×N"],
        ["Data raccolta", "2026-06-07"],
        ["", ""],
        ["Regola overcrowd", "G > N: floor(N²/G) carte/testa, minimo 3, resto nel mazzo"],
        ["Regola max giocatori", "G ≤ 2N"],
        ["Formato consigliato", "G = N (mazzo pesca 0)"],
        ["Riserva", "N carte scoperte dal tallone (N = lato griglia), condivise; posabili al posto della mano"],
        ["Cooperativo", "Universo noto via scheda 64 carte + dialogo (mani/mazzo coperti; riserva scoperta; ordine pesca ignoto)"],
        ["", ""],
        ["Copertura", "Tutte le configurazioni ammesse G=1..2N testate (1000 partite)"],
    ]
    for line in info:
        wi.append(line)
    wi["A1"].font = Font(bold=True, size=14, name="Arial")
    wi.column_dimensions["A"].width = 22
    wi.column_dimensions["B"].width = 70

    wb.save(OUT)
    print(f"Scritto: {OUT}")


if __name__ == "__main__":
    main()